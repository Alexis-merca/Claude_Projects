#!/usr/bin/env python3
"""Convertit Templates_diagnostic_Mercateam_v2.xlsx en JSON importable
dans Diagnostic OS (format `client_json`, version 1).

Un onglet use case = un processus. Les colonnes se posent ainsi :
    QUAND          -> etapes[].phase     (bandeau de frise du diagramme)
    ROLE           -> etapes[].role      (couloir)
    ACTION RELEVEE -> etapes[].texte
    SUPPORTS       -> etapes[].supports  (separateur VIRGULE, cf. listeSupports)
    ORDRE          -> position dans le tableau (l'importeur renumerote)
    N°             -> depart d'egalite uniquement
"""

import json
import re
import unicodedata

import openpyxl

SOURCE = ("/root/.claude/uploads/a0555995-c6c1-5fb2-8d4d-e4f6f9232744/"
          "8e6166de-Templates_diagnostic_Mercateam_v2.xlsx")
CIBLE = ("/tmp/claude-0/-home-user-Claude-Projects/"
         "a0555995-c6c1-5fb2-8d4d-e4f6f9232744/scratchpad/template-use-case.json")

# Onglets use case, dans l'ordre du classeur.
ONGLETS = [
    "Pilotage compétences",
    "Planification et gestion aléas",
    "Pilotage Charge capacité",
    "Intégration",
    "Transfert savoir-faire",
    "Standardisation sites",
    "Habilitations",
    "Audits",
    "Équité affectations",
    "Reconnaissance",
]

# Une virgule dans un nom de support le couperait en deux outils fantomes :
# `listeSupports` decoupe sur la virgule et rien d'autre.
SANS_VIRGULE = {
    "Logiciel (GED, ex VDOC)": "Logiciel (GED)",
    "Logiciel (autre, à préciser)": "Logiciel (autre)",
}


def net(v):
    return "" if v is None else re.sub(r"\s+", " ", str(v)).strip()


def slug(v):
    v = unicodedata.normalize("NFD", v.lower())
    v = "".join(c for c in v if unicodedata.category(c) != "Mn")
    return re.sub(r"(^-|-$)", "", re.sub(r"[^a-z0-9]+", "-", v))


def supports(brut):
    """« Papier + Mail » -> « Papier, Mail ». Deux outils, pas un."""
    brut = SANS_VIRGULE.get(net(brut), net(brut))
    morceaux = [SANS_VIRGULE.get(m.strip(), m.strip()) for m in brut.split("+")]
    return ", ".join(m for m in morceaux if m)


def entete(ws):
    """Ligne du bandeau de colonnes. Repere « N° » plutot qu'un index fige."""
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        if net(r[0]) == "N°":
            return i, [net(c) for c in r]
    raise SystemExit(f"{ws.title} : bandeau de colonnes introuvable")


wb = openpyxl.load_workbook(SOURCE, data_only=True)

# Ordre de reference des roles : celui de l'onglet « Référentiels ». Il place
# le collaborateur en haut et la direction en bas, et il est le meme pour les
# dix onglets — un role garde donc une hauteur comparable d'un processus a
# l'autre, ce que l'ordre d'apparition ne donnerait pas.
ref = wb["Référentiels"]
lg_ref, cols_ref = None, None
for i, r in enumerate(ref.iter_rows(values_only=True), 1):
    if "RÔLE" in [net(c) for c in r]:
        lg_ref, cols_ref = i, [net(c) for c in r]
        break
if lg_ref is None:
    raise SystemExit("Référentiels : bandeau de colonnes introuvable")
# Colonnes espacees d'une colonne vide : on lit les index, on ne les devine pas.
i_role, i_sup = cols_ref.index("RÔLE"), cols_ref.index("SUPPORTS")

ROLES_REF, SUPPORTS_REF = [], []
for r in ref.iter_rows(min_row=lg_ref + 1, values_only=True):
    if net(r[i_role]):
        ROLES_REF.append(net(r[i_role]))
    if net(r[i_sup]):
        SUPPORTS_REF.append(SANS_VIRGULE.get(net(r[i_sup]), net(r[i_sup])))

# Le menu deroulant ajoute un outil a la fois : les combinaisons « A + B » du
# referentiel s'obtiennent par deux clics, elles n'ont pas leur place ici.
OUTILS = [s for s in SUPPORTS_REF if "+" not in s]

processus, journal = [], []

for onglet in ONGLETS:
    ws = wb[onglet]
    titre = net(ws.cell(1, 1).value)
    commercial = net(ws.cell(2, 1).value)
    perimetre = re.sub(r"^Périmètre\s*:\s*", "", net(ws.cell(3, 1).value))

    lg, cols = entete(ws)
    col = {nom: i for i, nom in enumerate(cols)}

    brut = []
    for r in ws.iter_rows(min_row=lg + 1, values_only=True):
        action = net(r[col["ACTION RELEVÉE"]])
        if not action:
            continue
        brut.append({
            "num": float(net(r[col["N°"]]) or 0),
            "ordre": float(net(r[col["ORDRE"]]) or 0),
            "quand": net(r[col["QUAND"]]),
            "role": net(r[col["RÔLE"]]),
            "texte": action,
            "supports": supports(r[col["SUPPORTS"]]),
        })

    # ORDRE fait foi ; N° ne sert qu'a departager deux etapes paralleles.
    brut.sort(key=lambda e: (e["ordre"], e["num"]))

    utilises = []
    for e in brut:
        if e["role"] and e["role"] not in utilises:
            utilises.append(e["role"])
    roles = [r for r in ROLES_REF if r in utilises]
    hors_ref = [r for r in utilises if r not in ROLES_REF]
    roles += hors_ref
    if hors_ref:
        journal.append(f"{onglet} : role hors referentiel -> {hors_ref}")

    processus.append({
        "code": slug(onglet),
        "nom": titre,
        "soustitre": perimetre,
        "rang": len(processus) + 1,
        "roles": roles,
        "etapes": [{
            "ordre": i,
            "role": e["role"],
            "role2": "",
            "phase": e["quand"],
            "texte": e["texte"],
            "supports": e["supports"],
            "lien": "",
        } for i, e in enumerate(brut, 1)],
        # Le classeur ne porte ni friction ni chiffre : ils se relevent en
        # entretien. Les inventer viderait la trame de son sens.
        "frictions": [],
        "chiffres": [],
    })
    journal.append(f"{onglet} : {len(brut)} etapes, {len(roles)} roles | {commercial}")

garde = wb["Page de garde"]
correspondance = []
for r in garde.iter_rows(values_only=True):
    c = [net(x) for x in r]
    if c[0].startswith(("UC ", "Transv")) and net(r[1]):
        correspondance.append(f"  {c[0]} · {c[1]} · {c[3]}")

notes = "\n".join([
    "Trame de diagnostic Mercateam — un processus par use case.",
    "Source : Templates_diagnostic_Mercateam_v2.xlsx (version 2, huit logigrammes).",
    "",
    "Hypothèse de départ : ce relevé est la retranscription d'une routine d'usine",
    "lambda, peu outillée et peu mature (niveau 1 à 2 sur 5). C'est une base de",
    "discussion pour l'entretien, pas un constat. Chaque étape se confirme, se",
    "corrige, se réordonne ou se supprime avec l'interlocuteur. Une étape que",
    "l'interlocuteur ne reconnaît pas est soit à supprimer, soit le signe qu'il",
    "décrit la cible et non la pratique réelle : il faut alors creuser. Une étape",
    "qu'il trouve « pire que ça chez nous » est un irritant à noter tel quel, avec",
    "ses mots.",
    "",
    "Le bandeau de frise du diagramme porte le QUAND : moment ou fréquence réelle",
    "de l'action. C'est le meilleur détecteur d'irritant — une action quotidienne",
    "qui devrait être mensuelle est un point de douleur. Un support « Au jugé » ou",
    "« Oral » sur une étape de décision est un signal fort à remonter dans le",
    "rapport. Sur un site plus avancé, remplacer ces deux supports par l'outil",
    "réellement utilisé.",
    "",
    "UC 9 et UC 10 sont des use cases Advanced : sur un site de maturité 1 à 2, il",
    "n'y a souvent aucun process constitué. Les utiliser comme une liste de",
    "questions plutôt que comme un déroulé à confirmer.",
    "",
    "Correspondance use case / nom de version diagnostic / catégorie de valeur :",
    *correspondance,
])

sortie = {
    "format": "diagnostic-os",
    "version": 1,
    "client": {
        "code": "template-use-case",
        "nom": "Template use case",
        "site": "Trame de référence",
        "date_visite": "",
        "notes": notes,
        "outils": OUTILS,
        "si": {},
        "processus": processus,
    },
}

with open(CIBLE, "w", encoding="utf-8") as f:
    json.dump(sortie, f, ensure_ascii=False, indent=2)
    f.write("\n")

for l in journal:
    print(l)
print("---")
print("processus :", len(processus))
print("etapes    :", sum(len(p["etapes"]) for p in processus))
print("outils    :", len(OUTILS))
